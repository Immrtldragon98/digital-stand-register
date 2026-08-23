from calendar import monthrange
from datetime import datetime, timedelta
from io import BytesIO
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.stand_asset import StandAsset
from app.models.enums import LocationEnum
from app.models.entry_guide_asset import EntryGuideAsset
from app.models.line import Line
from app.models.stand_installation import StandInstallation


class ReportService:
    def __init__(self, db: Session):
        self.db = db

    def get_asset_inventory_summary(self):
        ready_stands = self.db.query(StandAsset).filter(StandAsset.current_location == LocationEnum.READY_AREA).all()
        wip_stands = self.db.query(StandAsset).filter(StandAsset.current_location == LocationEnum.WIP).all()
        ready_guides = self.db.query(EntryGuideAsset).filter(EntryGuideAsset.current_location == LocationEnum.READY_AREA).all()
        wip_guides = self.db.query(EntryGuideAsset).filter(EntryGuideAsset.current_location == LocationEnum.WIP).all()
        return {
            "ready_stands_count": len(ready_stands),
            "wip_stands_count": len(wip_stands),
            "ready_guides_count": len(ready_guides),
            "wip_guides_count": len(wip_guides),
            "ready_stands": [{"id": s.id, "code": s.code, "current_status": s.current_status} for s in ready_stands],
            "wip_stands": [{"id": s.id, "code": s.code, "current_status": s.current_status} for s in wip_stands],
            "ready_guides": [{"id": g.id, "code": g.code, "current_status": g.current_status} for g in ready_guides],
            "wip_guides": [{"id": g.id, "code": g.code, "current_status": g.current_status} for g in wip_guides],
        }

    def _stand_at(self, position_id: int, moment: datetime):
        inst = self.db.query(StandInstallation).filter(
            StandInstallation.position_id == position_id,
            StandInstallation.installed_at <= moment,
            (StandInstallation.removed_at.is_(None)) | (StandInstallation.removed_at > moment),
        ).order_by(StandInstallation.installed_at.desc()).first()
        return inst.stand.code if inst else ""

    def build_monthly_running_status_xlsx(self, year: int, month: int) -> BytesIO:
        if month < 1 or month > 12:
            raise ValueError("month must be between 1 and 12")

        wb = Workbook()
        wb.remove(wb.active)
        lines = self.db.query(Line).order_by(Line.id).all()
        days = monthrange(year, month)[1]
        today = datetime.utcnow().date()

        header_fill = PatternFill("solid", fgColor="2F2F2F")
        header_font = Font(color="FFFFFF", bold=True)

        for line in lines:
            ws = wb.create_sheet(title=(line.name or f"WRM-{line.id}")[:31])
            headers = ["Date"] + [f"Stand {i}" for i in range(1, 11)]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            positions = {p.position_number: p for p in line.positions}
            for day in range(1, days + 1):
                d = datetime(year, month, day).date()
                if d > today:
                    break
                # End-of-day snapshot gives the final running arrangement for that date.
                moment = datetime(year, month, day, 23, 59, 59)
                row = [d.strftime("%d-%m-%Y")]
                for pos_num in range(1, 11):
                    pos = positions.get(pos_num)
                    row.append(self._stand_at(pos.id, moment) if pos else "")
                ws.append(row)

            ws.freeze_panes = "A2"
            ws.column_dimensions["A"].width = 14
            for col in range(2, 12):
                ws.column_dimensions[chr(64 + col)].width = 12
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
